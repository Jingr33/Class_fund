import numpy as np
import customtkinter as ctk
import tkinter as tk
from student import Student
from payment import Payment
import os 
from pathlib import Path
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

class Logic ():
    """Programm logic."""
    def __init__(self):
        self.students = self._load_db()

##### STUDENTS DATABASE #####
    def _load_db (self) -> list:
        """Load data from database."""
        lines = []
        with open(self._get_resource_path("database.txt"), "r") as f:
            lines = f.readlines()
        
        students = []
        for i in range(len(lines)):
            data = lines[i].replace("\n", "").split(" ")
            student = Student(data[0], data[1], data[2], int(float(data[3])))
            students.append(student)
        return students
    
    def _add_student(self, first_name : str, surname : str):
        """Add student into the student list."""
        student = Student(len(self.students) + 1, first_name.capitalize(), surname.capitalize(), 0)
        self.students.append(student)
        self.students.sort(key = lambda student: student.surname)
        self._regenerate_ids()
        self._save_db()

    def _remove_student(self, id : int):
        """Remove student from the student list."""
        for student in self.students:
            if (student.id == id):
                self.students.remove(student)
        self._regenerate_ids()
        self._save_db()
    
    def _regenerate_ids (self) -> None:
        """Set new id for every student."""
        for i in range(len(self.students)):
            self.students[i].id = i + 1

    def _save_db (self):
        """Save students into the database."""
        with open(self._get_resource_path("database.txt"), "w") as f:
            for student in self.students:
                line = "{0} {1} {2} {3}\n".format(student.id, student.first_name, student.surname, student.account)
                f.write(line)

    def _choose_some(self, ids : list) -> list:
        """Choose some elected students from all students."""
        elected_students = []
        for student in self.students():
            if (student.id in ids):
                elected_students.append(student)
        return elected_students
        
    def choose_all_students (self, student_frames, value : bool):
        """Switch to all students selected or to noone stelected variant, if it is clicked to the select all checkbox."""
        if (value):
            self._set_students_marks(student_frames, True)
        else:
            self._set_students_marks(student_frames, False)

    def _set_students_marks(self, student_frames, value : bool) -> None:
        """Set all students attributes and checkboxes to entered value."""
        for student in self.students:
            student.choosen = value
        for frame in student_frames:
            frame.value.set(value)

    def _set_one_student_mark(self, elected_student : Student, value : bool) -> None:
        """Set a right mark value to elected student."""
        for student in self.students:
            if (student == elected_student):
                student.choosen = value

    def _get_marked_students (self) -> list:
        """Return a list of choosen students."""
        marked = []
        for student in self.students:
            if (student.choosen):
                marked.append(student)
        return marked

#### ACCOUNT ####
    def _total_amount(self, students : list) -> int:
        """Sum of total amout in the fund."""
        sum = 0
        for student in students:
            sum = sum + student.account
        return sum
    
    def _student_ave(self, sum : int, students : list) -> int:
        """Average of a one student from total amount."""
        if (len(students) == 0):
            return 0
        return np.ceil(sum / len(students))
    
    def _add_eachone_amount(self, amount : int, students : list) -> None:
        """Add an entered amount for each marked student."""
        for student in students:
            student.account = student.account + amount

    def _add_amount_from_sum(self, sum : int, students : list) -> None:
        """Add average of entred sum to each marked student."""
        if (len(students) == 0):
            return
        average = np.floor(sum / len(students))
        self._add_eachone_amount(average, students)

    def _remove_eachone_amount(self, amount : int, students : list) -> bool:
        """Check if it is possible to remove entered value and remove an entered amount for each marked student."""
        for student in students:
            if (student.account < amount): 
                return False
        for student in students:
            student.account = student.account - amount
        return True
    
    def _remove_amount_from_sum(self, sum : int, students : list) -> bool:
        """Check if it is possoble to remove average value from each student account a remove average of entred sum to each marked student."""
        if (len(students) == 0):
            return
        average = np.ceil(sum / len(students))
        checked = self._remove_eachone_amount(average, students)
        return checked

#### SAVE PAYMENTS #### 
    def _load_payments_db (self) -> list:
        """Load the payments_database data and return data as a list."""
        lines = []
        with open (self._get_resource_path("payment_database.txt"), "r") as f:
            lines = f.readlines()
        payments = []
        for line in lines:
            data = line.replace("\n", "").split(" ")
            payment = Payment(data[0], data[1], data[2], data[3], data[4:], self.students)
            payments.append(payment)
        return payments
        

    def _save_payment(self, name : str, value : int, students : list, type : str):
        """Save data about payment into payment database."""
        name = self._check_empty_str(name)
        today = datetime.now().date().strftime("%Y-%m-%d")
        sum, one = self._get_sum_and_value(value, students, type)
        student_names = ""
        for student in students:
            student_names = student_names + " " + student.surname
        payment = "{0} {1} {2} {3} {4}\n".format(name, today, sum, one, student_names)
        with open (self._get_resource_path("payment_database.txt"), "a") as pd:
            pd.write(payment)

    def _get_sum_and_value (self, value : int, students : list, type : str):
        """Return sum and value of one student from one payment."""
        if (len(students) == 0):
            return 0, 0
        elif (type == "add_sum"):
            return value, np.floor(value / len(students))
        elif (type == "remove_sum"):
            return -value, -np.ceil(value / len(students))
        elif (type == "one_add"):
            return len(students) * value, value
        else:
            return -len(students) * value, -value 
        
    def _check_empty_str(self, text : str) -> str:
        """Check if a string is empty, if yes, return text : nezadano."""
        if (text == ""):
            return "nezadáno"
        return text

#### FILE PATH ####
    def _get_resource_path(self, relative_path):
        """Find the right file path if the app runs in python and also in exe file"""
        try:
            # if the app runs in exe file
            base_path = sys._MEIPASS
        except AttributeError:
            # if the app runs in python
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)
    
### CREATE EXEL TABLE ####
    def create_table(self):
        """Create an excel table with all students, paymants and account."""
        wb = Workbook()
        ws = wb.active # + přejmenovat list
        ws.title = "Třídní fond"
        self._write_student_db(ws)
        self._write_payment_db(ws)
        self._change_cells_sizes(ws)
        self._style_table(ws)
        self._set_cells_fill(ws)
        wb.save(self._path_to_downloads("tridni_font.xlsx"))

    def _write_student_db (self, ws):
        """Write students data to a excel table from students database.txt."""
        students = self._load_db()
        ws.merge_cells("A1:D2")
        ws["A1"] = "Třídní fond"
        #content
        for i in range(len(students)):
            ws["A{0}".format(i+4)] = int(students[i].id)
            ws["B{0}".format(i+4)] = students[i].first_name
            ws["C{0}".format(i+4)] = students[i].surname
            ws["D{0}".format(i+4)] = students[i].account
        #legend
        ws["A3"] = "Pořadí"
        ws["B3"] = "Jméno"
        ws["C3"] = "Příjmení"
        ws["D3"] = "Konto"
        ws.merge_cells("A{0}:C{0}".format(len(self.students) + 4))
        ws["A{0}".format(len(self.students) + 4)] = "Celkem"
        ws["D{0}".format(len(self.students) + 4)] = self._total_amount(self.students)

    def _write_payment_db (self, ws) -> None:
        """Write payments data to the class fund excel."""
        self.payments = self._load_payments_db()
        if (len(self.payments) > 0):
            ws.merge_cells("F1:{0}1".format(get_column_letter(len(self.payments) + 5))) # title merge
        ws["F1"] = "Platby"
        column = 6
        for i in range(len(self.payments)-1, -1, -1):
            cell_letter = get_column_letter(column)
            ws["{0}2".format(cell_letter)] = self.payments[i].name
            ws["{0}3".format(cell_letter)] = self.payments[i].date
            self._write_students_payments(ws, self.payments[i], cell_letter)
            ws["{0}{1}".format(cell_letter, len(self.students) + 4)] = int(self.payments[i].sum_value)
            column = column + 1

    def _write_students_payments (self, ws, payment : Payment, cell_letter : str) -> None:
        """Write a value of each student value of one payment into the excel table."""
        for j in range(len(self.students)):
            if (self.students[j] in payment.who_payed and payment.who_payed[self.students[j]] == True):
                ws["{0}{1}".format(cell_letter, j+4)] = int(payment.one_value)

    def _change_cells_sizes(self, ws) -> None:
        """Update heights and widths od the rows and column for better design."""
        #widths
        ws.column_dimensions["A"].width = 7
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 7
        for j in range(len(self.payments)):
            ws.column_dimensions[get_column_letter(j+6)].width = 13.5

        #heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 30
        for i in range(len(self.students)):
            ws.row_dimensions[i+4].height = 20
        ws.row_dimensions[len(self.students) + 4].height = 30

    def _style_table(self, ws) -> None:
        """Set a table styling."""
        self._set_table_borders(ws)
        self._set_alignment(ws)
        self._set_letter_styles(ws)
        self._set_cells_fill(ws)
        
    def _set_table_borders (self, ws) -> None:
        thin_border = Border(
            left = Side(style="thin"),
            top = Side(style="thin"),
            right = Side(style="thin"),
            bottom = Side(style="thin")
        )
        for i in range(1, len(self.students) + 5):
            for j in range(1, len(self.payments) + 6):
                cell_letter = get_column_letter(j)
                if (cell_letter.upper() != "E"):
                    ws["{0}{1}".format(cell_letter, i)].border = thin_border

    def _set_alignment (self, ws) -> None:
        """Set alingment for cells."""
        # center center
        sum_row = len(self.students) + 4
        cc_cells = ["A1", "A3", "B3", "C3", "D3", "A{0}".format(sum_row), "D{0}".format(sum_row)]
        cc_alig = Alignment(horizontal="center", vertical="center")
        for cell in cc_cells:
            ws[cell].alignment = cc_alig
        for h in range(6, len(self.payments) + 6):
            ws["{0}{1}".format(get_column_letter(h), sum_row)].alignment = cc_alig
        # center bottom
        cb_alig = Alignment(horizontal="center", vertical="bottom")
        for i in range(4, len(self.students) + 4):
            ws["A{0}".format(i)].alignment = cb_alig
            ws["D{0}".format(i)].alignment = cb_alig
        for j in range(6, len(self.payments) + 6):
            col_letter = get_column_letter(j)
            ws["{0}2".format(col_letter)].alignment = cb_alig
            for k in range(4, len(self.students) + 4):
                ws["{0}{1}".format(col_letter, k)].alignment = cb_alig
        # left center
        ws["F1"].alignment = Alignment(vertical="center", horizontal="left")

    def _set_letter_styles (self, ws) -> None:
        """Set a letter styles. (font size and style)"""
        ws["A1"].font = Font(size = 15, bold = True)
        ws["F1"].font = Font(size = 13, bold = True)
        bold = Font(bold = True)
        sum_row = len(self.students) + 4
        for i in range(1, len(self.payments) + 6):
            col_letter = get_column_letter(i)
            if (col_letter.upper() != "E"):
                ws["{0}{1}".format(col_letter, sum_row)].font = bold # last row
                if (i < 5):
                    ws["{0}3".format(col_letter)].font = bold
                else:
                    ws["{0}2".format(col_letter)].font = bold
        for j in range(4, sum_row):
            ws["D{0}".format(j)].font = bold

    def _set_cells_fill(self, ws):
        """Set a fill colors for cells."""
        primary = PatternFill(start_color="7da3d1", end_color="7da3d1", fill_type="solid")
        secondary = PatternFill(start_color="b8d2f2", end_color="b8d2f2", fill_type="solid")
        tertiary = PatternFill(start_color="dce7f5", end_color="dce7f5", fill_type="solid")
        sum_row = len(self.students) + 4
        # primary fill
        primary_cells = ["A1", "F1", "A{0}".format(sum_row), "D{0}".format(sum_row)]
        for cell in primary_cells:
            ws[cell].fill = primary
        # secondary fill
        for i in range(1, len(self.payments) + 6):
            if (i == 5):
                continue
            elif (i < 5):
                ws["{0}3".format(get_column_letter(i))].fill = secondary
            else:
                ws["{0}2".format(get_column_letter(i))].fill = secondary
                ws["{0}{1}".format(get_column_letter(i), sum_row)].fill = secondary
        for j in range(4, sum_row):
            ws["D{0}".format(j)].fill = secondary
        # tertiary fill
        for k in range(4, sum_row):
            for l in ["A", "B", "C"]:
                ws["{0}{1}".format(l, k)].fill = tertiary
        for m in range(6, len(self.payments) + 6):
            ws["{0}3".format(get_column_letter(m))].fill = tertiary

    def _path_to_downloads(self, file_name : str) -> str:
        """Find and return path to downloads in user computer."""
        home_dir = Path.home()
        downloads_dir = home_dir / "Downloads"
        return os.path.join(downloads_dir, file_name)